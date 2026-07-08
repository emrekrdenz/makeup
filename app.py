#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import importlib.util
import json
import ssl
import sys
import threading
import time
import traceback
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple
from urllib.parse import unquote, urlparse

APP_ROOT = Path(__file__).resolve().parent
VENDOR = APP_ROOT / "vendor"
if VENDOR.exists():
    sys.path.insert(0, str(VENDOR))

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


OUTPUT_ROOT = APP_ROOT / "outputs"
RUNS_ROOT = OUTPUT_ROOT / "runs"
SCRAPER_PATH = OUTPUT_ROOT / "gratis_ingredients_scraper.py"
WATSONS_SCRAPER_PATH = OUTPUT_ROOT / "watsons_scraper.py"


def load_scraper():
    if not SCRAPER_PATH.exists():
        raise FileNotFoundError(f"Scraper bulunamadi: {SCRAPER_PATH}")

    spec = importlib.util.spec_from_file_location("gratis_ingredients_scraper", SCRAPER_PATH)
    if not spec or not spec.loader:
        raise RuntimeError("Scraper modulu yuklenemedi.")

    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


SCRAPER = load_scraper()


def load_watsons_scraper():
    if not WATSONS_SCRAPER_PATH.exists():
        return None

    spec = importlib.util.spec_from_file_location("watsons_scraper", WATSONS_SCRAPER_PATH)
    if not spec or not spec.loader:
        raise RuntimeError("Watsons scraper modulu yuklenemedi.")

    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


WATSONS = load_watsons_scraper()


@dataclass
class Job:
    id: str
    created_at: str
    status: str = "queued"
    message: str = "Sirada"
    total_products: int = 0
    processed_products: int = 0
    rows: int = 0
    images: int = 0
    ingredients_found: int = 0
    errors: List[Dict[str, str]] = field(default_factory=list)
    logs: List[str] = field(default_factory=list)
    run_dir: Optional[Path] = None
    csv_path: Optional[Path] = None
    manifest_path: Optional[Path] = None
    xlsx_path: Optional[Path] = None
    zip_path: Optional[Path] = None

    def log(self, message: str) -> None:
        stamp = datetime.now().strftime("%H:%M:%S")
        line = f"[{stamp}] {message}"
        self.logs.append(line)
        self.logs = self.logs[-500:]
        if self.run_dir:
            (self.run_dir / "job.log").write_text("\n".join(self.logs) + "\n", encoding="utf-8")


JOBS: Dict[str, Job] = {}
JOBS_LOCK = threading.Lock()


def parse_urls(value: str) -> List[str]:
    urls: List[str] = []
    seen = set()
    for raw in value.replace(",", "\n").splitlines():
        url = raw.strip()
        if not url:
            continue
        if not urlparse(url).scheme:
            url = "https://" + url
        if url not in seen:
            seen.add(url)
            urls.append(url)
    return urls


def site_for_url(url: str) -> str:
    host = urlparse(url).netloc.lower()
    if "watsons.com.tr" in host:
        return "watsons"
    return "gratis"


def int_option(payload: Dict[str, object], key: str, default: int, minimum: int, maximum: int) -> int:
    try:
        value = int(payload.get(key, default))
    except (TypeError, ValueError):
        value = default
    return max(minimum, min(maximum, value))


def bool_option(payload: Dict[str, object], key: str, default: bool = False) -> bool:
    value = payload.get(key, default)
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in {"1", "true", "yes", "on"}
    return bool(value)


def collect_gratis_targets(
    category_url: str,
    context: ssl.SSLContext,
    timeout: int,
    retries: int,
    page_delay: float,
    page_limit: Optional[int],
    keep_duplicates: bool,
    job: Job,
) -> Tuple[List[object], int, int]:
    first_source = SCRAPER.fetch_html(
        SCRAPER.category_page_url(category_url, 1),
        context,
        timeout,
        retries,
        page_delay,
    )
    max_page = SCRAPER.extract_max_page(first_source)
    if page_limit:
        max_page = min(max_page, page_limit)

    occurrences: List[Tuple[str, int, int]] = []
    pages_by_url: Dict[str, set] = {}
    first_seen_by_url: Dict[str, int] = {}

    for page in range(1, max_page + 1):
        source = first_source if page == 1 else SCRAPER.fetch_html(
            SCRAPER.category_page_url(category_url, page),
            context,
            timeout,
            retries,
            page_delay,
        )
        links = SCRAPER.extract_product_links(source, category_url)
        job.log(f"{category_url} | sayfa {page}/{max_page}: {len(links)} urun linki")

        for position, url in enumerate(links, start=1):
            occurrences.append((url, page, position))
            pages_by_url.setdefault(url, set()).add(page)
            first_seen_by_url.setdefault(url, len(first_seen_by_url) + 1)

    if keep_duplicates:
        targets = [
            SCRAPER.ProductTarget(url=url, pages=(page,), first_seen=index)
            for index, (url, page, _position) in enumerate(occurrences, start=1)
        ]
    else:
        targets = [
            SCRAPER.ProductTarget(url=url, pages=tuple(sorted(pages)), first_seen=first_seen_by_url[url])
            for url, pages in pages_by_url.items()
        ]
        targets.sort(key=lambda item: item.first_seen)

    return targets, len(occurrences), max_page


def collect_targets_for_site(
    site: str,
    category_url: str,
    context: ssl.SSLContext,
    timeout: int,
    retries: int,
    page_limit: Optional[int],
    keep_duplicates: bool,
    job: Job,
) -> Tuple[List[object], int, int]:
    if site == "watsons":
        if WATSONS is None:
            raise RuntimeError("Watsons scraper bulunamadi.")
        return WATSONS.collect_product_targets(
            category_url=category_url,
            timeout=timeout,
            retries=retries,
            page_delay=0.15,
            page_limit=page_limit,
            keep_duplicates=keep_duplicates,
            log=job.log,
        )

    return collect_gratis_targets(
        category_url=category_url,
        context=context,
        timeout=timeout,
        retries=retries,
        page_delay=0.15,
        page_limit=page_limit,
        keep_duplicates=keep_duplicates,
        job=job,
    )


def parse_product_for_site(
    site: str,
    target: object,
    category_url: str,
    context: ssl.SSLContext,
    timeout: int,
    retries: int,
    download_images: bool,
    image_root: Optional[Path],
    max_images: int,
) -> Dict[str, str]:
    if site == "watsons":
        if WATSONS is None:
            raise RuntimeError("Watsons scraper bulunamadi.")
        return WATSONS.parse_product(
            target=target,
            category_url=category_url,
            timeout=timeout,
            retries=retries,
            product_delay=0.1,
            download_images=download_images,
            image_root=image_root,
            image_delay=0.02,
            max_images_per_product=max_images,
        )

    return SCRAPER.parse_product(
        target,
        category_url,
        context,
        timeout,
        retries,
        0.1,
        download_images,
        image_root,
        0.02,
        max_images,
    )


def split_cell_list(value: str) -> List[str]:
    return [item.strip() for item in (value or "").split(" | ") if item.strip()]


def relative_to_run(path_value: str, run_dir: Path) -> str:
    if not path_value:
        return ""
    try:
        return str(Path(path_value).resolve().relative_to(run_dir.resolve()))
    except Exception:
        return path_value


def enrich_image_columns(
    rows: Sequence[Dict[str, str]],
    run_dir: Path,
) -> List[Dict[str, str]]:
    enriched_rows: List[Dict[str, str]] = []

    for row_number, row in enumerate(rows, start=1):
        image_files = split_cell_list(row.get("image_files", ""))
        image_urls = split_cell_list(row.get("image_urls", ""))
        image_file = image_files[0] if image_files else ""
        image_url = image_urls[0] if image_urls else ""
        new_row = dict(row)

        new_row["row_number"] = str(row_number)
        new_row["image_count"] = "1" if image_file or image_url else "0"
        new_row["image_file"] = relative_to_run(image_file, run_dir) if image_file else ""
        new_row["image_source_path"] = image_file
        new_row["image_url"] = image_url

        enriched_rows.append(new_row)

    return enriched_rows


def write_dict_csv(rows: Sequence[Dict[str, str]], output_path: Path, fieldnames: Sequence[str]) -> None:
    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_combined_csv(rows: Sequence[Dict[str, str]], output_path: Path) -> None:
    fieldnames = [
        "row_number",
        "source_category",
        "category_pages",
        "product_id",
        "title",
        "barcode",
        "ingredients_found",
        "ingredients",
        "image_count",
        "image_file",
        "image_url",
        "url",
    ]
    write_dict_csv(rows, output_path, fieldnames)


def make_link(cell, target: str) -> None:
    if not target:
        return
    cell.value = target
    cell.hyperlink = target
    cell.style = "Hyperlink"


def write_products_xlsx(rows: Sequence[Dict[str, str]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "products"
    preview_dir = output_path.parent / "_xlsx_previews"
    preview_dir.mkdir(parents=True, exist_ok=True)

    headers = [
        "row_number",
        "preview",
        "product_id",
        "barcode",
        "title",
        "ingredients_found",
        "ingredients",
        "image_count",
        "image_file",
        "image_url",
        "product_url",
        "source_category",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="E8F4F1")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    widths = {
        "A": 12,
        "B": 18,
        "C": 14,
        "D": 18,
        "E": 44,
        "F": 16,
        "G": 70,
        "H": 13,
        "I": 48,
        "J": 48,
        "K": 48,
        "L": 48,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    for row_index, row in enumerate(rows, start=2):
        values = [
            row.get("row_number", ""),
            "",
            row.get("product_id", ""),
            row.get("barcode", ""),
            row.get("title", ""),
            row.get("ingredients_found", ""),
            row.get("ingredients", ""),
            row.get("image_count", ""),
            row.get("image_file", ""),
            row.get("image_url", ""),
            row.get("url", ""),
            row.get("source_category", ""),
        ]
        sheet.append(values)
        sheet.row_dimensions[row_index].height = 86
        for col_index in range(1, len(headers) + 1):
            sheet.cell(row=row_index, column=col_index).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        sheet.cell(row=row_index, column=9).value = row.get("image_file", "")
        if row.get("image_source_path"):
            sheet.cell(row=row_index, column=9).hyperlink = row.get("image_source_path", "")
            sheet.cell(row=row_index, column=9).style = "Hyperlink"
        make_link(sheet.cell(row=row_index, column=10), row.get("image_url", ""))
        make_link(sheet.cell(row=row_index, column=11), row.get("url", ""))
        make_link(sheet.cell(row=row_index, column=12), row.get("source_category", ""))

        image_path = row.get("image_source_path", "")
        if image_path and Path(image_path).exists():
            try:
                preview_path = preview_dir / f"preview_{row.get('row_number', row_index)}.png"
                with PILImage.open(image_path) as source_image:
                    source_image.thumbnail((180, 180))
                    if source_image.mode not in ("RGB", "RGBA"):
                        source_image = source_image.convert("RGBA")
                    source_image.save(preview_path, "PNG")
                image = ExcelImage(str(preview_path))
                image.width = 76
                image.height = 76
                sheet.add_image(image, f"B{row_index}")
            except Exception:
                sheet.cell(row=row_index, column=2).value = "preview yok"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_path)


def zip_directory(source_dir: Path, zip_path: Path) -> int:
    count = 0
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(source_dir.rglob("*")):
            if path.is_file():
                archive.write(path, path.relative_to(source_dir.parent))
                count += 1
    return count


def run_job(job_id: str, payload: Dict[str, object]) -> None:
    with JOBS_LOCK:
        job = JOBS[job_id]
        job.status = "running"
        job.message = "Basladi"

    urls = parse_urls(str(payload.get("urls", "")))
    download_images = bool_option(payload, "downloadImages", True)
    insecure = bool_option(payload, "insecure", True)
    keep_duplicates = bool_option(payload, "keepDuplicates", False)
    workers = int_option(payload, "workers", 4, 1, 12)
    timeout = int_option(payload, "timeout", 25, 5, 120)
    retries = int_option(payload, "retries", 2, 0, 5)
    page_limit_raw = int_option(payload, "pageLimit", 0, 0, 500)
    product_limit = int_option(payload, "productLimit", 0, 0, 100000)
    max_images = 1

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = RUNS_ROOT / f"{stamp}_{job_id[:8]}"
    image_root = run_dir / "images"
    csv_path = run_dir / "products.csv"
    xlsx_path = run_dir / "products.xlsx"
    zip_path = run_dir / "images.zip"
    run_dir.mkdir(parents=True, exist_ok=True)

    with JOBS_LOCK:
        job.run_dir = run_dir
        job.csv_path = csv_path
        job.manifest_path = None
        job.xlsx_path = xlsx_path
        job.zip_path = zip_path if download_images else None
        job.log(f"Job klasoru: {run_dir}")

    try:
        if not urls:
            raise ValueError("URL listesi bos.")

        context = SCRAPER.build_ssl_context(insecure)
        all_targets: List[Tuple[str, str, object]] = []
        total_cards = 0

        for category_url in urls:
            site = site_for_url(category_url)
            targets, occurrence_count, max_page = collect_targets_for_site(
                site=site,
                category_url=category_url,
                context=context,
                timeout=timeout,
                retries=retries,
                page_limit=page_limit_raw or None,
                keep_duplicates=keep_duplicates,
                job=job,
            )
            total_cards += occurrence_count
            all_targets.extend((site, category_url, target) for target in targets)
            job.log(
                f"{category_url} | site: {site}, kart: {occurrence_count}, islenecek: {len(targets)}, sayfa: {max_page}"
            )
            if occurrence_count != len(targets):
                job.log(
                    f"{category_url} | {occurrence_count - len(targets)} tekrar eden kart tekillestirildi"
                )

        if product_limit:
            all_targets = all_targets[:product_limit]

        with JOBS_LOCK:
            job.total_products = len(all_targets)
            job.message = f"{len(all_targets)} urun isleniyor"

        rows: List[Dict[str, str]] = []
        image_root.mkdir(parents=True, exist_ok=True)

        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_map = {
                executor.submit(
                    parse_product_for_site,
                    site,
                    target,
                    category_url,
                    context,
                    timeout,
                    retries,
                    download_images,
                    image_root if download_images else None,
                    max_images,
                ): (site, category_url, target)
                for site, category_url, target in all_targets
            }

            for future in as_completed(future_map):
                site, category_url, target = future_map[future]
                try:
                    row = future.result()
                    rows.append(row)
                    with JOBS_LOCK:
                        job.processed_products += 1
                        job.rows = len(rows)
                        job.images += int(row.get("image_count") or 0)
                        if row.get("ingredients_found") == "yes":
                            job.ingredients_found += 1
                        job.message = f"{job.processed_products}/{job.total_products} urun"
                    job.log(f"OK {target.url}")
                except Exception as exc:
                    with JOBS_LOCK:
                        job.processed_products += 1
                        job.errors.append({"url": target.url, "error": str(exc)})
                        job.message = f"{job.processed_products}/{job.total_products} urun"
                    job.log(f"HATA {target.url}: {exc}")

        order_by_url = {target.url: index for index, (_site, _category_url, target) in enumerate(all_targets)}
        rows.sort(key=lambda row: order_by_url.get(row["url"], 10**9))

        enriched_rows = enrich_image_columns(rows, run_dir)
        write_combined_csv(enriched_rows, csv_path)
        write_products_xlsx(enriched_rows, xlsx_path)
        job.log(f"CSV yazildi: {csv_path}")
        job.log(f"Excel yazildi: {xlsx_path}")

        if download_images:
            zipped_count = zip_directory(image_root, zip_path)
            job.log(f"Gorsel ZIP yazildi: {zip_path} ({zipped_count} dosya)")

        with JOBS_LOCK:
            job.status = "done" if not job.errors else "done_with_errors"
            job.message = f"Bitti: {len(rows)} urun, {job.images} gorsel"

    except Exception as exc:
        trace = traceback.format_exc()
        with JOBS_LOCK:
            job.status = "failed"
            job.message = str(exc)
            job.errors.append({"url": "", "error": str(exc)})
            job.log(trace)


def job_snapshot(job: Job) -> Dict[str, object]:
    downloads = []
    if job.csv_path and job.csv_path.exists():
        downloads.append({"label": "Ürün CSV", "url": f"/download/{job.id}/products.csv"})
    if job.xlsx_path and job.xlsx_path.exists():
        downloads.append({"label": "Excel", "url": f"/download/{job.id}/products.xlsx"})
    if job.zip_path and job.zip_path.exists():
        downloads.append({"label": "Gorseller ZIP", "url": f"/download/{job.id}/images.zip"})
    if job.run_dir and (job.run_dir / "job.log").exists():
        downloads.append({"label": "Log", "url": f"/download/{job.id}/job.log"})

    return {
        "id": job.id,
        "createdAt": job.created_at,
        "status": job.status,
        "message": job.message,
        "totalProducts": job.total_products,
        "processedProducts": job.processed_products,
        "rows": job.rows,
        "images": job.images,
        "ingredientsFound": job.ingredients_found,
        "errors": job.errors[-25:],
        "logs": job.logs[-120:],
        "downloads": downloads,
    }


INDEX_HTML = r"""<!doctype html>
<html lang="tr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>URL Ürün Toplayıcı</title>
  <style>
    :root {
      --bg: #f5f7f8;
      --panel: #ffffff;
      --text: #172022;
      --muted: #627074;
      --line: #d7e0e2;
      --accent: #00856f;
      --accent-dark: #006b5a;
      --warn: #a65f00;
      --bad: #b42318;
      --good: #067647;
      --soft: #e8f4f1;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background: var(--bg);
      color: var(--text);
      font-family: ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      font-size: 14px;
      letter-spacing: 0;
    }
    main {
      width: min(1180px, calc(100% - 32px));
      margin: 0 auto;
      padding: 24px 0 36px;
    }
    .topbar {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      margin-bottom: 16px;
    }
    h1 {
      margin: 0;
      font-size: 24px;
      line-height: 1.2;
      font-weight: 760;
    }
    .status-pill {
      min-width: 128px;
      padding: 8px 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--panel);
      color: var(--muted);
      text-align: center;
      font-weight: 650;
    }
    .layout {
      display: grid;
      grid-template-columns: minmax(0, 1.05fr) minmax(360px, .95fr);
      gap: 16px;
      align-items: start;
    }
    section {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
    }
    h2 {
      margin: 0 0 12px;
      font-size: 15px;
      line-height: 1.3;
    }
    label {
      display: block;
      margin: 0 0 6px;
      color: var(--muted);
      font-weight: 650;
      font-size: 12px;
    }
    .hint {
      margin-top: 6px;
      color: var(--warn);
      font-size: 12px;
      font-weight: 700;
      line-height: 1.35;
    }
    textarea, input {
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #fff;
      color: var(--text);
      font: inherit;
      outline: none;
    }
    textarea {
      min-height: 220px;
      resize: vertical;
      padding: 12px;
      line-height: 1.45;
    }
    input {
      height: 38px;
      padding: 8px 10px;
    }
    textarea:focus, input:focus { border-color: var(--accent); box-shadow: 0 0 0 3px var(--soft); }
    .grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 12px;
      margin-top: 14px;
    }
    .checks {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px 12px;
      margin-top: 14px;
    }
    .check {
      display: flex;
      align-items: center;
      min-height: 36px;
      gap: 8px;
      color: var(--text);
      font-weight: 650;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 8px 10px;
      background: #fbfcfc;
    }
    .check input {
      width: 16px;
      height: 16px;
      padding: 0;
      accent-color: var(--accent);
    }
    .check-text {
      display: flex;
      flex-direction: column;
      gap: 2px;
      line-height: 1.25;
    }
    .check-text small {
      color: var(--muted);
      font-size: 11px;
      font-weight: 650;
    }
    .actions {
      display: flex;
      gap: 10px;
      margin-top: 16px;
      flex-wrap: wrap;
    }
    button {
      min-height: 40px;
      border: 0;
      border-radius: 8px;
      padding: 0 16px;
      color: #fff;
      background: var(--accent);
      cursor: pointer;
      font-weight: 750;
    }
    button:hover { background: var(--accent-dark); }
    button.secondary {
      background: #eef3f4;
      color: var(--text);
      border: 1px solid var(--line);
    }
    button.secondary:hover { background: #e3ebed; }
    button:disabled {
      opacity: .55;
      cursor: not-allowed;
    }
    .stats {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 8px;
      margin-bottom: 12px;
    }
    .metric {
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
      min-height: 72px;
      background: #fbfcfc;
    }
    .metric strong {
      display: block;
      font-size: 22px;
      line-height: 1;
      margin-bottom: 8px;
    }
    .metric span { color: var(--muted); font-size: 12px; font-weight: 650; }
    .progress {
      height: 10px;
      border-radius: 999px;
      background: #e6ecee;
      overflow: hidden;
      border: 1px solid #d8e1e3;
      margin: 4px 0 14px;
    }
    .bar {
      height: 100%;
      width: 0;
      background: var(--accent);
      transition: width .25s ease;
    }
    .downloads {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      min-height: 40px;
      margin-bottom: 12px;
    }
    .downloads a {
      display: inline-flex;
      align-items: center;
      min-height: 36px;
      padding: 0 12px;
      border-radius: 8px;
      color: var(--accent-dark);
      background: var(--soft);
      text-decoration: none;
      font-weight: 760;
    }
    pre {
      margin: 0;
      height: 330px;
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 12px;
      background: #111819;
      color: #d7f8ed;
      font-size: 12px;
      line-height: 1.45;
      white-space: pre-wrap;
    }
    .errors {
      margin-top: 12px;
      color: var(--bad);
      line-height: 1.45;
      max-height: 150px;
      overflow: auto;
    }
    @media (max-width: 860px) {
      main { width: min(100% - 20px, 760px); padding-top: 14px; }
      .layout { grid-template-columns: 1fr; }
      .stats { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .grid, .checks { grid-template-columns: 1fr; }
      .topbar { align-items: flex-start; flex-direction: column; }
      .status-pill { width: 100%; }
    }
  </style>
</head>
<body>
<main>
  <div class="topbar">
    <h1>URL Ürün Toplayıcı</h1>
    <div class="status-pill" id="status">Hazır</div>
  </div>

  <div class="layout">
    <section>
      <h2>Kaynaklar</h2>
      <label for="urls">URL listesi</label>
      <textarea id="urls" spellcheck="false">https://www.watsons.com.tr/makyaj/c/100</textarea>
      <div class="hint">Watsons kategori URL'si genelde /c/ kodu içerir. Örn: /makyaj/c/100</div>

      <div class="grid">
        <div>
          <label for="workers">Paralel ürün</label>
          <input id="workers" type="number" min="1" max="12" value="4">
        </div>
        <div>
          <label for="pageLimit">Sayfa limiti</label>
          <input id="pageLimit" type="number" min="0" max="500" value="0">
          <div class="hint">0 = sınırsız</div>
        </div>
        <div>
          <label for="productLimit">Ürün limiti</label>
          <input id="productLimit" type="number" min="0" max="100000" value="0">
          <div class="hint">0 = sınırsız</div>
        </div>
      </div>

      <div class="checks">
        <label class="check"><input id="downloadImages" type="checkbox" checked> Görseller</label>
        <label class="check"><input id="insecure" type="checkbox" checked> SSL esnek</label>
        <label class="check">
          <input id="keepDuplicates" type="checkbox">
          <span class="check-text">
            <span>Tekrar eden ürünleri de yaz</span>
            <small>Kapalıysa aynı ürün URL'si tekilleştirilir</small>
          </span>
        </label>
      </div>

      <div class="actions">
        <button id="startBtn">Başlat</button>
        <button class="secondary" id="watsonsSampleBtn" type="button">Watsons Örnek</button>
        <button class="secondary" id="gratisSampleBtn" type="button">Gratis Örnek</button>
      </div>
    </section>

    <section>
      <h2>Durum</h2>
      <div class="stats">
        <div class="metric"><strong id="mProducts">0</strong><span>Ürün</span></div>
        <div class="metric"><strong id="mDone">0</strong><span>İşlenen</span></div>
        <div class="metric"><strong id="mImages">0</strong><span>Görsel</span></div>
        <div class="metric"><strong id="mIngredients">0</strong><span>İçerik</span></div>
      </div>
      <div class="progress"><div class="bar" id="bar"></div></div>
      <div class="downloads" id="downloads"></div>
      <pre id="log"></pre>
      <div class="errors" id="errors"></div>
    </section>
  </div>
</main>

<script>
const $ = (id) => document.getElementById(id);
let currentJob = null;
let timer = null;

function payload() {
  return {
    urls: $("urls").value,
    workers: Number($("workers").value || 4),
    pageLimit: Number($("pageLimit").value || 0),
    productLimit: Number($("productLimit").value || 0),
    downloadImages: $("downloadImages").checked,
    insecure: $("insecure").checked,
    keepDuplicates: $("keepDuplicates").checked
  };
}

function statusText(status) {
  const map = {
    queued: "Sırada",
    running: "Çalışıyor",
    done: "Bitti",
    done_with_errors: "Bitti",
    failed: "Hata"
  };
  return map[status] || "Hazır";
}

function render(job) {
  $("status").textContent = statusText(job.status);
  $("mProducts").textContent = job.totalProducts || 0;
  $("mDone").textContent = job.processedProducts || 0;
  $("mImages").textContent = job.images || 0;
  $("mIngredients").textContent = job.ingredientsFound || 0;
  const pct = job.totalProducts ? Math.round((job.processedProducts / job.totalProducts) * 100) : 0;
  $("bar").style.width = `${Math.min(100, pct)}%`;
  $("log").textContent = (job.logs || []).join("\n");
  $("log").scrollTop = $("log").scrollHeight;
  $("downloads").innerHTML = (job.downloads || []).map(d => `<a href="${d.url}">${d.label}</a>`).join("");
  $("errors").innerHTML = (job.errors || []).map(e => `<div>${e.url ? e.url + " | " : ""}${e.error}</div>`).join("");
  $("startBtn").disabled = job.status === "queued" || job.status === "running";
}

async function poll() {
  if (!currentJob) return;
  const res = await fetch(`/api/jobs/${currentJob}`);
  const job = await res.json();
  render(job);
  if (["done", "done_with_errors", "failed"].includes(job.status)) {
    clearInterval(timer);
    timer = null;
    $("startBtn").disabled = false;
  }
}

$("startBtn").addEventListener("click", async () => {
  $("startBtn").disabled = true;
  $("log").textContent = "";
  $("errors").innerHTML = "";
  $("downloads").innerHTML = "";
  const res = await fetch("/api/jobs", {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify(payload())
  });
  const job = await res.json();
  currentJob = job.id;
  render(job);
  if (timer) clearInterval(timer);
  timer = setInterval(poll, 1200);
  poll();
});

$("watsonsSampleBtn").addEventListener("click", () => {
  $("urls").value = "https://www.watsons.com.tr/makyaj/c/100";
  $("pageLimit").value = 1;
  $("productLimit").value = 5;
});

$("gratisSampleBtn").addEventListener("click", () => {
  $("urls").value = "https://www.gratis.com/sac-bakim/sac-kremleri-c-50302";
  $("pageLimit").value = 1;
  $("productLimit").value = 5;
});
</script>
</body>
</html>
"""


class AppHandler(BaseHTTPRequestHandler):
    server_version = "WatsonsScraper/1.0"

    def log_message(self, format: str, *args) -> None:
        return

    def send_json(self, payload: Dict[str, object], status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_text(self, body: str, content_type: str = "text/html; charset=utf-8") -> None:
        data = body.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/":
            self.send_text(INDEX_HTML)
            return

        if parsed.path.startswith("/api/jobs/"):
            job_id = parsed.path.rsplit("/", 1)[-1]
            with JOBS_LOCK:
                job = JOBS.get(job_id)
                if not job:
                    self.send_json({"error": "Job bulunamadi."}, 404)
                    return
                self.send_json(job_snapshot(job))
            return

        if parsed.path.startswith("/download/"):
            self.handle_download(parsed.path)
            return

        self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path != "/api/jobs":
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length).decode("utf-8")
        payload = json.loads(raw or "{}")

        job_id = uuid.uuid4().hex
        job = Job(id=job_id, created_at=datetime.now().isoformat(timespec="seconds"))
        with JOBS_LOCK:
            JOBS[job_id] = job

        thread = threading.Thread(target=run_job, args=(job_id, payload), daemon=True)
        thread.start()

        self.send_json(job_snapshot(job), 202)

    def handle_download(self, path: str) -> None:
        parts = [unquote(part) for part in path.split("/") if part]
        if len(parts) != 3:
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        _download, job_id, filename = parts
        with JOBS_LOCK:
            job = JOBS.get(job_id)
            if not job or not job.run_dir:
                self.send_error(HTTPStatus.NOT_FOUND)
                return
            candidate = (job.run_dir / filename).resolve()

        if candidate.parent != job.run_dir.resolve() or not candidate.exists():
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        content_type = "application/octet-stream"
        if candidate.suffix == ".csv":
            content_type = "text/csv; charset=utf-8"
        elif candidate.suffix == ".xlsx":
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif candidate.suffix == ".zip":
            content_type = "application/zip"
        elif candidate.suffix == ".log":
            content_type = "text/plain; charset=utf-8"

        data = candidate.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{candidate.name}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def main() -> int:
    RUNS_ROOT.mkdir(parents=True, exist_ok=True)
    host = "127.0.0.1"
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8765
    server = ThreadingHTTPServer((host, port), AppHandler)
    print(f"URL Ürün Toplayıcı: http://{host}:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nKapatiliyor.")
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
